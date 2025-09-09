
Módulo de lógica (API).
Lee el archivo Excel, filtra la información y calcula las medianas.


from openpyxl import load_workbook
import statistics

# Ruta del archivo Excel
RUTA_EXCEL = r"C:\Users\pc\OneDrive\Escritorio\Progra4\Parcial1\main\data\resultado_laboratorio_suelo.xlsx"

def consultar_datos(departamento, municipio, cultivo, n_registros):
    try:
        # Abrir el archivo Excel directamente
        wb = load_workbook(RUTA_EXCEL)
        hoja = wb.active  # Usa la primera hoja

        # Leer encabezados (primera fila)
        headers = [cell.value for cell in hoja[1]]

        # Identificar índices de columnas importantes (usando los nombres reales del Excel)
        idx_dep = headers.index("Departamento")
        idx_mun = headers.index("Municipio")
        idx_cul = headers.index("Cultivo")
        idx_top = headers.index("Topografia")
        idx_ph  = headers.index("pH agua:suelo 2,5:1,0")
        idx_p   = headers.index("Fósforo (P) Bray II mg/kg")
        idx_k   = headers.index("Potasio (K) intercambiable cmol(+)/kg")

        # Filtrar filas según los criterios
        registros = []
        for row in hoja.iter_rows(min_row=2, values_only=True):
            if (row[idx_dep] == departamento and 
                row[idx_mun] == municipio and 
                row[idx_cul] == cultivo):
                registros.append(row)

        if not registros:
            return []

        # Limitar al número de registros solicitado
        registros = registros[:n_registros]

        # Extraer columnas para calcular medianas
        ph_vals = [r[idx_ph] for r in registros if isinstance(r[idx_ph], (int, float))]
        p_vals  = [r[idx_p]  for r in registros if isinstance(r[idx_p], (int, float))]
        k_vals  = [r[idx_k]  for r in registros if isinstance(r[idx_k], (int, float))]

        # Extraer todas las topografías distintas
        topografias = list({r[idx_top] for r in registros if r[idx_top] is not None})

        medianas = {
            "Departamento": departamento,
            "Municipio": municipio,
            "Cultivo": cultivo,
            "Topografias": topografias,
            "Mediana_pH": statistics.median(ph_vals) if ph_vals else None,
            "Mediana_P": statistics.median(p_vals) if p_vals else None,
            "Mediana_K": statistics.median(k_vals) if k_vals else None
        }

        return [medianas]

    except Exception as e:
        print("Error al consultar los datos:", e)
        return []

